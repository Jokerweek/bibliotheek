from openpyxl import load_workbook
import json

def excel_to_json(excel_file_path, json_file_path):
    # Load the workbook and select the active sheet
    workbook = load_workbook(filename=excel_file_path)
    sheet = workbook.active

    # Iterate through rows in the sheet and construct JSON objects
    json_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        json_obj = {}
        tags = []
        for i, cell in enumerate(row):
            column_name = sheet.cell(row=1, column=i+1).value
            # Combine tag1, tag2, tag3 into a list of tags, excluding None values
            if column_name in ['tag1', 'tag2', 'tag3']:
                if cell:  # Ensure cell is not None
                    tags.append(cell)
            else:
                json_obj[column_name] = cell
        json_obj['tags'] = tags
        json_list.append(json_obj)

    # Convert the list of dictionaries (JSON objects) to a JSON string and save it
    with open(json_file_path, 'w', encoding='utf-8') as file:
        json.dump(json_list, file, ensure_ascii=False, indent=2)

# Specify the path to your Excel file and the desired JSON output file path
excel_file_path = 'path/to/your/excel.xlsx'  # Change this to your actual Excel file path
json_file_path = 'path/to/your/output.json'  # Change this to your desired output JSON file path

# Convert the Excel file to JSON
excel_to_json(excel_file_path, json_file_path)

print(f'JSON data has been saved to {json_file_path}')
